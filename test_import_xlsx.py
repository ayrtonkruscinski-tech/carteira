#!/usr/bin/env python3

import requests
import tempfile
import os
import json
from openpyxl import Workbook

def test_xlsx_import_with_dates():
    """Test XLSX import with purchase_date extraction"""
    
    base_url = "https://stockmaster-178.preview.emergentagent.com/api"
    headers = {'Authorization': 'Bearer demo_session_token_123'}
    
    print("🔍 Testing XLSX Import with Purchase Date Extraction...")
    
    # Step 1: Clear existing stocks
    print("   🗑️  Clearing existing stocks...")
    response = requests.delete(f"{base_url}/portfolio/stocks/all", headers=headers, timeout=10)
    print(f"   Clear response: {response.status_code}")
    
    # Step 2: Create test XLSX with different date formats
    wb = Workbook()
    ws = wb.active
    
    # Headers
    ws['A1'] = 'ticker'
    ws['B1'] = 'name'
    ws['C1'] = 'quantity'
    ws['D1'] = 'average_price'
    ws['E1'] = 'purchase_date'
    ws['F1'] = 'sector'
    
    # Data rows with different date formats
    ws['A2'] = 'PETR4'
    ws['B2'] = 'Petrobras PN'
    ws['C2'] = 100
    ws['D2'] = 35.50
    ws['E2'] = '15/01/2024'  # DD/MM/YYYY format
    ws['F2'] = 'Petróleo'
    
    ws['A3'] = 'VALE3'
    ws['B3'] = 'Vale ON'
    ws['C3'] = 50
    ws['D3'] = 62.30
    ws['E3'] = '2024-02-20'  # YYYY-MM-DD format
    ws['F3'] = 'Mineração'
    
    ws['A4'] = 'ITUB4'
    ws['B4'] = 'Itaú Unibanco PN'
    ws['C4'] = 200
    ws['D4'] = 32.80
    ws['E4'] = '10/03/2024'  # DD/MM/YYYY format
    ws['F4'] = 'Bancos'
    
    ws['A5'] = 'PETR4'  # Duplicate for aggregation test
    ws['B5'] = 'Petrobras PN'
    ws['C5'] = 50
    ws['D5'] = 40.00
    ws['E5'] = '20/01/2024'  # Later date - should not be kept
    ws['F5'] = 'Petróleo'
    
    # Step 3: Save and upload XLSX
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        wb.save(f.name)
        temp_file_path = f.name
    
    try:
        with open(temp_file_path, 'rb') as f:
            files = {'file': ('test_portfolio.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            
            print(f"   📤 Uploading XLSX to: {base_url}/portfolio/import")
            response = requests.post(f"{base_url}/portfolio/import", files=files, headers=headers, timeout=30)
            
            print(f"   📊 Import response: {response.status_code}")
            if response.status_code == 200:
                result = response.json()
                print(f"   ✅ Import result: {result}")
            else:
                print(f"   ❌ Import failed: {response.text}")
                return False
            
        # Step 4: Check imported stocks
        print("   📊 Checking imported stocks...")
        response = requests.get(f"{base_url}/portfolio/stocks", headers=headers, timeout=10)
        
        if response.status_code == 200:
            stocks = response.json()
            print(f"   📊 Found {len(stocks)} stocks:")
            
            for stock in stocks:
                ticker = stock.get('ticker')
                purchase_date = stock.get('purchase_date')
                quantity = stock.get('quantity')
                
                print(f"     {ticker}: quantity={quantity}, purchase_date='{purchase_date}'")
                
                # Check date format
                if purchase_date:
                    if len(purchase_date) == 10 and purchase_date[4] == '-' and purchase_date[7] == '-':
                        print(f"       ✅ Correct YYYY-MM-DD format")
                    else:
                        print(f"       ❌ Incorrect format (expected YYYY-MM-DD)")
                else:
                    print(f"       ❌ No purchase_date")
            
            # Check PETR4 aggregation
            petr4_stocks = [s for s in stocks if s.get('ticker') == 'PETR4']
            if len(petr4_stocks) == 1:
                petr4 = petr4_stocks[0]
                print(f"   📊 PETR4 aggregation: quantity={petr4.get('quantity')}, date={petr4.get('purchase_date')}")
                
                if petr4.get('quantity') == 150:
                    print(f"     ✅ Quantity correctly aggregated (100+50=150)")
                else:
                    print(f"     ❌ Quantity not aggregated correctly")
                    
                # Check if earliest date was kept (15/01/2024 should be earlier than 20/01/2024)
                expected_date = "2024-01-15"  # 15/01/2024 converted
                if petr4.get('purchase_date') == expected_date:
                    print(f"     ✅ Earliest date kept: {expected_date}")
                else:
                    print(f"     ❌ Earliest date not kept. Got: {petr4.get('purchase_date')}, expected: {expected_date}")
            else:
                print(f"   ❌ Expected 1 PETR4 stock, found {len(petr4_stocks)}")
                
        else:
            print(f"   ❌ Failed to get stocks: {response.status_code} - {response.text}")
            return False
            
    finally:
        # Clean up
        if os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
    
    return True

if __name__ == "__main__":
    test_xlsx_import_with_dates()