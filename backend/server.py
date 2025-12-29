from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import csv
import io
import re
from openpyxl import load_workbook
from tradingview_ta import TA_Handler, Interval
from bs4 import BeautifulSoup
import requests

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Alpha Vantage config
ALPHA_VANTAGE_KEY = os.environ.get('ALPHA_VANTAGE_KEY', 'demo')
ALPHA_VANTAGE_BASE = "https://www.alphavantage.co/query"

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Portfolio(BaseModel):
    portfolio_id: str = Field(default_factory=lambda: f"port_{uuid.uuid4().hex[:12]}")
    user_id: str
    name: str
    description: Optional[str] = None
    is_default: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PortfolioCreate(BaseModel):
    name: str
    description: Optional[str] = None

class PortfolioUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

class UserSession(BaseModel):
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Stock(BaseModel):
    stock_id: str = Field(default_factory=lambda: f"stock_{uuid.uuid4().hex[:12]}")
    user_id: str
    portfolio_id: Optional[str] = None  # ID da carteira
    ticker: str
    name: str
    quantity: float
    average_price: float
    purchase_date: Optional[str] = None
    sector: Optional[str] = None
    current_price: Optional[float] = None
    dividend_yield: Optional[float] = None
    ceiling_price: Optional[float] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StockCreate(BaseModel):
    ticker: str
    name: str
    quantity: float
    average_price: float
    purchase_date: Optional[str] = None
    sector: Optional[str] = None
    current_price: Optional[float] = None
    dividend_yield: Optional[float] = None
    ceiling_price: Optional[float] = None
    portfolio_id: Optional[str] = None  # ID da carteira

class StockUpdate(BaseModel):
    quantity: Optional[float] = None
    average_price: Optional[float] = None
    purchase_date: Optional[str] = None
    current_price: Optional[float] = None
    dividend_yield: Optional[float] = None
    ceiling_price: Optional[float] = None

class Dividend(BaseModel):
    dividend_id: str = Field(default_factory=lambda: f"div_{uuid.uuid4().hex[:12]}")
    user_id: str
    stock_id: str
    ticker: str
    portfolio_id: Optional[str] = None  # ID da carteira
    amount: float
    payment_date: str
    type: str = "dividendo"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DividendCreate(BaseModel):
    stock_id: str
    ticker: str
    amount: float
    payment_date: str
    type: str = "dividendo"
    portfolio_id: Optional[str] = None  # ID da carteira

class PortfolioSnapshot(BaseModel):
    snapshot_id: str = Field(default_factory=lambda: f"snap_{uuid.uuid4().hex[:12]}")
    user_id: str
    date: str
    total_invested: float
    total_current: float
    total_dividends: float
    stocks_count: int
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Alert(BaseModel):
    alert_id: str = Field(default_factory=lambda: f"alert_{uuid.uuid4().hex[:12]}")
    user_id: str
    stock_id: str
    ticker: str
    alert_type: str  # "ceiling_reached", "price_drop", "price_rise"
    message: str
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ValuationRequest(BaseModel):
    ticker: str
    current_price: float
    dividend_per_share: float
    dividend_growth_rate: float = 0.05
    discount_rate: float = 0.12
    desired_yield: float = 0.06
    free_cash_flow: Optional[float] = None
    shares_outstanding: Optional[float] = None
    growth_rate: Optional[float] = 0.05
    # Campos para método Warren Buffett
    net_income: Optional[float] = None  # Lucro Líquido
    depreciation: Optional[float] = None  # Depreciação e Amortização
    capex: Optional[float] = None  # Capital Expenditure
    roe: Optional[float] = None  # Return on Equity (%)
    payout: Optional[float] = None  # Payout ratio (%)

class AnalysisRequest(BaseModel):
    ticker: str
    current_price: float
    sector: Optional[str] = None
    dividend_yield: Optional[float] = None
    pe_ratio: Optional[float] = None
    question: Optional[str] = None

# ==================== AUTH HELPERS ====================

async def get_current_user(request: Request) -> User:
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session_doc = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    
    return User(**user_doc)

# ==================== AUTH ROUTES ====================

@api_router.get("/auth/session")
async def exchange_session(request: Request, response: Response):
    session_id = request.headers.get("X-Session-ID")
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing session ID")
    
    async with httpx.AsyncClient() as http_client:
        resp = await http_client.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id}
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session")
        data = resp.json()
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    existing_user = await db.users.find_one({"email": data["email"]}, {"_id": 0})
    
    if existing_user:
        user_id = existing_user["user_id"]
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": data["name"], "picture": data.get("picture")}}
        )
    else:
        await db.users.insert_one({
            "user_id": user_id,
            "email": data["email"],
            "name": data["name"],
            "picture": data.get("picture"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    session_token = data.get("session_token", f"sess_{uuid.uuid4().hex}")
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    await db.user_sessions.delete_many({"user_id": user_id})
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return user_doc

@api_router.get("/auth/me")
async def get_me(user: User = Depends(get_current_user)):
    return user.model_dump()

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logged out"}

# ==================== INVESTIDOR10 SCRAPER ====================

def fetch_investidor10_fundamentals(ticker: str) -> dict:
    """Fetch fundamental data from Investidor10 for valuation"""
    data = {
        "ticker": ticker.upper(),
        "current_price": None,
        "dividend_per_share": None,
        "dividend_yield": None,
        "p_l": None,  # P/L ratio
        "p_vp": None,  # P/VP ratio
        "roe": None,
        "payout": None,  # Payout ratio
        "lpa": None,  # Lucro por Ação
        "vpa": None,  # Valor Patrimonial por Ação
        "net_income": None,  # Lucro Líquido
        "ebitda": None,
        "net_revenue": None,
        "free_cash_flow": None,
        "shares_outstanding": None,
        "dividend_growth_rate": 5.0,  # Default
        "depreciation": None,
        "capex": None,
        "market_cap": None,  # Valor de mercado
        "stock_type": "PN" if ticker.upper().endswith("4") else "ON",  # ON (3) ou PN (4)
    }
    
    try:
        # Use base ticker for company data (remove 3/4 suffix for some searches)
        base_ticker = ticker.upper()
        url = f"https://investidor10.com.br/acoes/{base_ticker.lower()}/"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=15)
        
        if response.status_code != 200:
            logger.error(f"Investidor10 fundamentals returned status {response.status_code} for {ticker}")
            return data
        
        soup = BeautifulSoup(response.content, 'lxml')
        
        # Helper function to parse Brazilian numbers
        def parse_br_number(text):
            if not text:
                return None
            text = text.strip().replace('R$', '').replace('%', '').strip()
            # Handle trillions/billions/millions
            multiplier = 1
            if 'trilh' in text.lower():
                multiplier = 1000000000000
                text = re.sub(r'trilh[ãa]o|trilh[õo]es?', '', text, flags=re.IGNORECASE).strip()
            elif 'bilh' in text.lower() or text.lower().endswith(' bi'):
                multiplier = 1000000000
                text = re.sub(r'bilh[ãa]o|bilh[õo]es?|bi$', '', text, flags=re.IGNORECASE).strip()
            elif 'milh' in text.lower() or text.lower().endswith(' mi'):
                multiplier = 1000000
                text = re.sub(r'milh[ãa]o|milh[õo]es?|mi$', '', text, flags=re.IGNORECASE).strip()
            elif 'mil' in text.lower():
                multiplier = 1000
                text = text.lower().replace('mil', '').strip()
            
            # Parse the number - handle Brazilian format: 1.234,56
            text = text.strip()
            if ',' in text and '.' in text:
                if text.rfind(',') > text.rfind('.'):
                    text = text.replace('.', '').replace(',', '.')
                else:
                    text = text.replace(',', '')
            elif ',' in text:
                text = text.replace(',', '.')
            
            try:
                return float(text) * multiplier
            except:
                return None
        
        page_text = soup.get_text()
        
        # Extract shares outstanding - look for "Nº total de papeis" pattern
        # Must be before market_cap calculation
        shares_patterns = [
            r'N[ºo°]\s*total\s*de\s*papeis?\s*([\d.,]+)\s*(Bilh[ãõo]es?|Milh[ãõo]es?|Trilh[ãõo]es?)',
            r'total\s*de\s*papeis?\s*([\d.,]+)\s*(Bilh[ãõo]es?|Milh[ãõo]es?|Trilh[ãõo]es?)',
        ]
        
        for pattern in shares_patterns:
            shares_match = re.search(pattern, page_text, re.IGNORECASE)
            if shares_match:
                shares_text = f"{shares_match.group(1)} {shares_match.group(2)}"
                data['shares_outstanding'] = parse_br_number(shares_text)
                logger.info(f"Found shares_outstanding for {ticker}: {data['shares_outstanding']} from pattern: {shares_text}")
                if data['shares_outstanding']:
                    break
        
        # Extract Net Income from "lucro no valor de R$ X Bilhões"
        lucro_match = re.search(r'lucro\s+(?:no\s+valor\s+de\s+)?R\$\s*([\d.,]+)\s*(Bilh[ãõo]es?|Milh[ãõo]es?|Trilh[ãõo]es?)?', page_text, re.IGNORECASE)
        if lucro_match:
            lucro_text = f"{lucro_match.group(1)} {lucro_match.group(2) or ''}"
            data['net_income'] = parse_br_number(lucro_text)
        
        # Extract Market Cap from "valor de mercado de R$ X Bilhões"
        market_match = re.search(r'valor de mercado de\s+R\$\s*([\d.,]+)\s*(Bilh[ãõo]es?|Milh[ãõo]es?|Trilh[ãõo]es?)?', page_text, re.IGNORECASE)
        if market_match:
            market_text = f"{market_match.group(1)} {market_match.group(2) or ''}"
            data['market_cap'] = parse_br_number(market_text)
        
        # Try to get ticker_id from script for API call
        ticker_id = None
        for script in soup.find_all('script'):
            if script.string:
                match = re.search(r'/api/historico-indicadores/(\d+)/', script.string)
                if match:
                    ticker_id = match.group(1)
                    break
        
        # Get indicators from API if ticker_id found
        if ticker_id:
            api_url = f'https://investidor10.com.br/api/historico-indicadores/{ticker_id}/10?v=2'
            api_response = requests.get(api_url, headers=headers, timeout=10)
            if api_response.ok:
                api_data = api_response.json()
                
                # Extract current values from API (index 0 = current/TTM)
                if 'LPA' in api_data and api_data['LPA']:
                    data['lpa'] = api_data['LPA'][0].get('value')
                if 'VPA' in api_data and api_data['VPA']:
                    data['vpa'] = api_data['VPA'][0].get('value')
                if 'P/L' in api_data and api_data['P/L']:
                    data['p_l'] = api_data['P/L'][0].get('value')
                if 'P/VP' in api_data and api_data['P/VP']:
                    data['p_vp'] = api_data['P/VP'][0].get('value')
                if 'DIVIDEND YIELD (DY)' in api_data and api_data['DIVIDEND YIELD (DY)']:
                    data['dividend_yield'] = api_data['DIVIDEND YIELD (DY)'][0].get('value')
                if 'MARGEM EBITDA' in api_data and api_data['MARGEM EBITDA']:
                    data['ebitda'] = api_data['MARGEM EBITDA'][0].get('value')
                
                # Calculate ROE average (last 5 YEARS - skip index 0 which is current/TTM)
                # Index 0 = atual/TTM, Index 1-5 = últimos 5 anos
                if 'ROE' in api_data and len(api_data['ROE']) > 1:
                    # Skip first item (current), get next 5 historical years
                    historical_roe = api_data['ROE'][1:6]
                    roe_values = [item.get('value') for item in historical_roe if item.get('value') is not None and item.get('value') > 0]
                    if roe_values:
                        data['roe'] = round(sum(roe_values) / len(roe_values), 2)
                        data['roe_years'] = len(roe_values)
                    data['roe_current'] = api_data['ROE'][0].get('value')
                
                # Calculate Payout average (last 5 YEARS - skip index 0 which is current/TTM)
                if 'PAYOUT' in api_data and len(api_data['PAYOUT']) > 1:
                    # Skip first item (current), get next 5 historical years
                    historical_payout = api_data['PAYOUT'][1:6]
                    # Filter outliers: payout should be between 0-100% normally, but allow up to 120% for some cases
                    payout_values = [item.get('value') for item in historical_payout if item.get('value') is not None and 0 < item.get('value') <= 120]
                    if payout_values:
                        data['payout'] = round(sum(payout_values) / len(payout_values), 2)
                        data['payout_years'] = len(payout_values)
                    data['payout_current'] = api_data['PAYOUT'][0].get('value')
        
        # Calculate shares_outstanding from market_cap and current_price if not found
        if not data['shares_outstanding'] and data['market_cap'] and data['current_price'] and data['current_price'] > 0:
            data['shares_outstanding'] = data['market_cap'] / data['current_price']
        
        # Calculate net_income from LPA and shares if not found
        if not data['net_income'] and data['lpa'] and data['shares_outstanding']:
            data['net_income'] = data['lpa'] * data['shares_outstanding']
        
        # Calculate growth rate from ROE and Payout for Buffett method
        # Growth = ROE × (1 - Payout)
        if data['roe'] and data['payout']:
            retention_rate = 1 - (data['payout'] / 100)
            if retention_rate > 0:
                data['growth_rate'] = data['roe'] * retention_rate / 100  # Convert to decimal
                data['dividend_growth_rate'] = data['growth_rate'] * 100  # Store as percentage
        
        # Estimate depreciation and capex based on typical ratios if not available
        if data['net_income'] and not data['depreciation']:
            data['depreciation'] = data['net_income'] * 0.15
        
        if data['net_income'] and not data['capex']:
            data['capex'] = data['net_income'] * 0.10
        
        # Calculate Owner Earnings (Buffett method)
        if data['net_income'] and data['depreciation'] is not None and data['capex'] is not None:
            data['free_cash_flow'] = data['net_income'] + data['depreciation'] - data['capex']
        
        logger.info(f"Investidor10 fundamentals for {ticker}: price={data['current_price']}, shares={data['shares_outstanding']}, net_income={data['net_income']}, roe={data['roe']}, payout={data['payout']}")
        
    except Exception as e:
        logger.error(f"Investidor10 fundamentals scraper error for {ticker}: {e}")
    
    return data

def fetch_investidor10_dividends(ticker: str) -> List[dict]:
    """Fetch dividend history from Investidor10"""
    dividends = []
    
    try:
        url = f"https://investidor10.com.br/acoes/{ticker.lower()}/"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=15)
        
        if response.status_code != 200:
            logger.error(f"Investidor10 returned status {response.status_code} for {ticker}")
            return []
        
        soup = BeautifulSoup(response.content, 'lxml')
        
        # Find dividend table - look for table with "data com" header
        tables = soup.find_all('table')
        
        for table in tables:
            headers_row = table.find('tr')
            if not headers_row:
                continue
                
            headers_text = [th.get_text(strip=True).lower() for th in headers_row.find_all(['th', 'td'])]
            
            # Check if this is the dividend table
            if 'data com' in headers_text or 'data ex' in headers_text:
                # Find column indices
                tipo_idx = next((i for i, h in enumerate(headers_text) if 'tipo' in h), None)
                data_com_idx = next((i for i, h in enumerate(headers_text) if 'data com' in h or 'data ex' in h), None)
                pagamento_idx = next((i for i, h in enumerate(headers_text) if 'pagamento' in h or 'data pag' in h), None)
                valor_idx = next((i for i, h in enumerate(headers_text) if 'valor' in h), None)
                
                if data_com_idx is None or valor_idx is None:
                    continue
                
                # Parse rows (skip header)
                rows = table.find_all('tr')[1:]
                
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) <= max(filter(None, [tipo_idx, data_com_idx, pagamento_idx, valor_idx])):
                        continue
                    
                    try:
                        tipo = cells[tipo_idx].get_text(strip=True) if tipo_idx is not None else "Dividendo"
                        data_com_str = cells[data_com_idx].get_text(strip=True) if data_com_idx is not None else ""
                        pagamento_str = cells[pagamento_idx].get_text(strip=True) if pagamento_idx is not None else ""
                        valor_str = cells[valor_idx].get_text(strip=True) if valor_idx is not None else "0"
                        
                        # Skip provisioned dividends (not yet confirmed)
                        if 'provisionado' in pagamento_str.lower() or 'provisionado' in tipo.lower():
                            logger.debug(f"Skipping provisioned dividend for {ticker}: {tipo}")
                            continue
                        
                        # Parse date (DD/MM/YYYY -> YYYY-MM-DD)
                        data_com = None
                        if data_com_str and '/' in data_com_str:
                            parts = data_com_str.split('/')
                            if len(parts) == 3:
                                data_com = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        
                        data_pagamento = None
                        if pagamento_str and '/' in pagamento_str:
                            parts = pagamento_str.split('/')
                            if len(parts) == 3:
                                data_pagamento = f"{parts[2]}-{parts[1]}-{parts[0]}"
                        
                        # Parse value (handle Brazilian format)
                        valor_str = valor_str.replace('.', '').replace(',', '.')
                        valor = float(valor_str) if valor_str else 0
                        
                        if data_com and valor > 0:
                            dividends.append({
                                "tipo": tipo,
                                "data_com": data_com,
                                "data_pagamento": data_pagamento,
                                "valor": round(valor, 8)
                            })
                    except Exception as e:
                        logger.error(f"Error parsing dividend row: {e}")
                        continue
                
                break  # Found the dividend table, no need to continue
        
        logger.info(f"Investidor10: Found {len(dividends)} dividends for {ticker}")
        
    except Exception as e:
        logger.error(f"Investidor10 scraper error for {ticker}: {e}")
    
    return dividends

# ==================== TRADINGVIEW INTEGRATION ====================

def fetch_tradingview_quote(ticker: str) -> dict:
    """Fetch real-time quote from TradingView"""
    try:
        handler = TA_Handler(
            symbol=ticker,
            screener="brazil",
            exchange="BMFBOVESPA",
            interval=Interval.INTERVAL_1_DAY
        )
        analysis = handler.get_analysis()
        
        indicators = analysis.indicators
        close_price = indicators.get("close", 0)
        open_price = indicators.get("open", 0)
        high_price = indicators.get("high", 0)
        low_price = indicators.get("low", 0)
        volume = indicators.get("volume", 0)
        change = indicators.get("change", 0)
        change_percent = indicators.get("change", 0)
        
        # Calculate change if not available
        if close_price and open_price and not change:
            change = close_price - open_price
            change_percent = ((close_price - open_price) / open_price) * 100 if open_price else 0
        
        return {
            "ticker": ticker,
            "price": round(close_price, 2) if close_price else 0,
            "open": round(open_price, 2) if open_price else 0,
            "high": round(high_price, 2) if high_price else 0,
            "low": round(low_price, 2) if low_price else 0,
            "change": round(change, 2) if change else 0,
            "change_percent": round(change_percent, 2) if change_percent else 0,
            "volume": int(volume) if volume else 0,
            "recommendation": analysis.summary.get("RECOMMENDATION", "NEUTRAL"),
            "source": "tradingview"
        }
    except Exception as e:
        logger.error(f"TradingView error for {ticker}: {e}")
    
    return None

# ==================== YAHOO FINANCE INTEGRATION (PRIMARY) ====================

async def fetch_yahoo_finance_quote(ticker: str) -> dict:
    """Fetch real-time quote from Yahoo Finance - PRIMARY SOURCE"""
    # Yahoo Finance uses .SA suffix for Brazilian stocks
    yahoo_ticker = f"{ticker}.SA"
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_ticker}"
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        async with httpx.AsyncClient() as http_client:
            resp = await http_client.get(url, headers=headers, timeout=10.0)
            data = resp.json()
            
            if 'chart' in data and 'result' in data['chart'] and data['chart']['result']:
                result = data['chart']['result'][0]
                meta = result.get('meta', {})
                
                price = meta.get('regularMarketPrice', 0)
                previous_close = meta.get('previousClose', 0)
                
                if price and price > 0:
                    change = price - previous_close if previous_close else 0
                    change_percent = (change / previous_close * 100) if previous_close else 0
                    
                    return {
                        "ticker": ticker,
                        "price": round(price, 2),
                        "change": round(change, 2),
                        "change_percent": round(change_percent, 2),
                        "previous_close": previous_close,
                        "source": "yahoo_finance"
                    }
                    
    except Exception as e:
        logger.error(f"Yahoo Finance error for {ticker}: {e}")
    
    return None

# ==================== ALPHA VANTAGE INTEGRATION (BACKUP) ====================

async def fetch_alpha_vantage_quote(ticker: str) -> dict:
    """Fetch real-time quote from Alpha Vantage"""
    # Convert Brazilian ticker to Alpha Vantage format (add .SAO suffix)
    av_ticker = f"{ticker}.SAO"
    
    try:
        async with httpx.AsyncClient() as http_client:
            resp = await http_client.get(
                ALPHA_VANTAGE_BASE,
                params={
                    "function": "GLOBAL_QUOTE",
                    "symbol": av_ticker,
                    "apikey": ALPHA_VANTAGE_KEY
                },
                timeout=10.0
            )
            data = resp.json()
            
            if "Global Quote" in data and data["Global Quote"]:
                quote = data["Global Quote"]
                return {
                    "ticker": ticker,
                    "price": float(quote.get("05. price", 0)),
                    "change": float(quote.get("09. change", 0)),
                    "change_percent": quote.get("10. change percent", "0%").replace("%", ""),
                    "volume": int(quote.get("06. volume", 0)),
                    "latest_trading_day": quote.get("07. latest trading day", ""),
                    "source": "alpha_vantage"
                }
    except Exception as e:
        logger.error(f"Alpha Vantage error for {ticker}: {e}")
    
    return None

# Brazilian stocks fallback data
BRAZILIAN_STOCKS = {
    "PETR4": {"ticker": "PETR4", "name": "Petrobras PN", "sector": "Petróleo", "current_price": 38.50, "dividend_yield": 12.5},
    "VALE3": {"ticker": "VALE3", "name": "Vale ON", "sector": "Mineração", "current_price": 62.30, "dividend_yield": 8.2},
    "ITUB4": {"ticker": "ITUB4", "name": "Itaú Unibanco PN", "sector": "Bancos", "current_price": 32.80, "dividend_yield": 5.1},
    "BBDC4": {"ticker": "BBDC4", "name": "Bradesco PN", "sector": "Bancos", "current_price": 14.20, "dividend_yield": 4.8},
    "BBAS3": {"ticker": "BBAS3", "name": "Banco do Brasil ON", "sector": "Bancos", "current_price": 28.90, "dividend_yield": 9.3},
    "WEGE3": {"ticker": "WEGE3", "name": "WEG ON", "sector": "Bens Industriais", "current_price": 42.50, "dividend_yield": 1.2},
    "RENT3": {"ticker": "RENT3", "name": "Localiza ON", "sector": "Consumo", "current_price": 45.60, "dividend_yield": 2.1},
    "MGLU3": {"ticker": "MGLU3", "name": "Magazine Luiza ON", "sector": "Varejo", "current_price": 2.15, "dividend_yield": 0.0},
    "ABEV3": {"ticker": "ABEV3", "name": "Ambev ON", "sector": "Bebidas", "current_price": 12.80, "dividend_yield": 5.5},
    "EGIE3": {"ticker": "EGIE3", "name": "Engie Brasil ON", "sector": "Energia", "current_price": 43.20, "dividend_yield": 7.8},
    "TAEE11": {"ticker": "TAEE11", "name": "Taesa Unit", "sector": "Energia", "current_price": 35.40, "dividend_yield": 9.5},
    "BBSE3": {"ticker": "BBSE3", "name": "BB Seguridade ON", "sector": "Seguros", "current_price": 35.20, "dividend_yield": 8.0},
    "SUZB3": {"ticker": "SUZB3", "name": "Suzano ON", "sector": "Papel e Celulose", "current_price": 58.90, "dividend_yield": 3.2},
    "JBSS3": {"ticker": "JBSS3", "name": "JBS ON", "sector": "Alimentos", "current_price": 34.50, "dividend_yield": 4.5},
    "HAPV3": {"ticker": "HAPV3", "name": "Hapvida ON", "sector": "Saúde", "current_price": 4.20, "dividend_yield": 0.0},
    "RADL3": {"ticker": "RADL3", "name": "Raia Drogasil ON", "sector": "Varejo", "current_price": 26.80, "dividend_yield": 0.8},
    "KLBN11": {"ticker": "KLBN11", "name": "Klabin Unit", "sector": "Papel e Celulose", "current_price": 22.40, "dividend_yield": 5.5},
    "CSAN3": {"ticker": "CSAN3", "name": "Cosan ON", "sector": "Energia", "current_price": 12.50, "dividend_yield": 6.2},
    "CPFE3": {"ticker": "CPFE3", "name": "CPFL Energia ON", "sector": "Energia", "current_price": 34.80, "dividend_yield": 8.5},
    "EMBR3": {"ticker": "EMBR3", "name": "Embraer ON", "sector": "Bens Industriais", "current_price": 52.30, "dividend_yield": 0.5},
}

@api_router.get("/stocks/search/{ticker}")
async def search_stock(ticker: str):
    ticker_upper = ticker.upper()
    
    # Try TradingView first (most reliable for Brazilian stocks)
    tv_data = fetch_tradingview_quote(ticker_upper)
    if tv_data and tv_data["price"] > 0:
        base_info = BRAZILIAN_STOCKS.get(ticker_upper, {})
        return {
            "ticker": ticker_upper,
            "name": base_info.get("name", f"Ação {ticker_upper}"),
            "sector": base_info.get("sector", "Outros"),
            "current_price": tv_data["price"],
            "dividend_yield": base_info.get("dividend_yield"),
            "change": tv_data["change"],
            "change_percent": tv_data["change_percent"],
            "high": tv_data.get("high"),
            "low": tv_data.get("low"),
            "volume": tv_data.get("volume"),
            "recommendation": tv_data.get("recommendation"),
            "source": "tradingview"
        }
    
    # Fallback to Alpha Vantage
    av_data = await fetch_alpha_vantage_quote(ticker_upper)
    if av_data and av_data["price"] > 0:
        base_info = BRAZILIAN_STOCKS.get(ticker_upper, {})
        return {
            "ticker": ticker_upper,
            "name": base_info.get("name", f"Ação {ticker_upper}"),
            "sector": base_info.get("sector", "Outros"),
            "current_price": av_data["price"],
            "dividend_yield": base_info.get("dividend_yield"),
            "change": av_data["change"],
            "change_percent": av_data["change_percent"],
            "source": "alpha_vantage"
        }
    
    # Fallback to cache data
    if ticker_upper in BRAZILIAN_STOCKS:
        return {**BRAZILIAN_STOCKS[ticker_upper], "source": "cache"}
    
    return {
        "ticker": ticker_upper,
        "name": f"Ação {ticker_upper}",
        "sector": "Outros",
        "current_price": None,
        "dividend_yield": None,
        "source": "unknown"
    }

@api_router.get("/stocks/quote/{ticker}")
async def get_stock_quote(ticker: str):
    """Get real-time quote for a stock from TradingView"""
    ticker_upper = ticker.upper()
    
    # Try Yahoo Finance first (most reliable)
    yahoo_data = await fetch_yahoo_finance_quote(ticker_upper)
    if yahoo_data and yahoo_data["price"] > 0:
        return yahoo_data
    
    # Fallback to TradingView
    tv_data = fetch_tradingview_quote(ticker_upper)
    if tv_data and tv_data["price"] > 0:
        return tv_data
    
    # Fallback to Alpha Vantage
    av_data = await fetch_alpha_vantage_quote(ticker_upper)
    if av_data:
        return av_data
    
    # Fallback to cache
    if ticker_upper in BRAZILIAN_STOCKS:
        return {
            "ticker": ticker_upper,
            "price": BRAZILIAN_STOCKS[ticker_upper]["current_price"],
            "change": 0,
            "change_percent": "0",
            "source": "cache"
        }
    
    raise HTTPException(status_code=404, detail="Stock not found")

@api_router.get("/stocks/valuation-data/{ticker}")
async def get_valuation_data(ticker: str):
    """Get fundamental data for valuation from Investidor10 and other sources"""
    ticker_upper = ticker.upper()
    
    # Get fundamentals from Investidor10
    fundamentals = fetch_investidor10_fundamentals(ticker_upper)
    
    # Get current price from Yahoo Finance (most reliable)
    yahoo_data = await fetch_yahoo_finance_quote(ticker_upper)
    if yahoo_data and yahoo_data["price"] > 0:
        fundamentals["current_price"] = yahoo_data["price"]
    
    # Get dividend info from dividend history
    dividends = fetch_investidor10_dividends(ticker_upper)
    if dividends:
        # Calculate annual dividend (sum of dividends from last 12 months by date)
        from datetime import datetime, timedelta
        one_year_ago = datetime.now() - timedelta(days=365)
        
        annual_dividend = 0
        for d in dividends:
            try:
                # data_com is in YYYY-MM-DD format
                dividend_date = datetime.strptime(d["data_com"], "%Y-%m-%d")
                if dividend_date >= one_year_ago:
                    annual_dividend += d["valor"]
            except (ValueError, KeyError):
                continue
        
        fundamentals["dividend_per_share"] = round(annual_dividend, 2)
        
        # Calculate dividend yield
        if fundamentals["current_price"] and fundamentals["current_price"] > 0:
            fundamentals["dividend_yield"] = round((annual_dividend / fundamentals["current_price"]) * 100, 2)
    
    # Add base info
    base_info = BRAZILIAN_STOCKS.get(ticker_upper, {})
    fundamentals["name"] = base_info.get("name", f"Ação {ticker_upper}")
    fundamentals["sector"] = base_info.get("sector", "Outros")
    
    return fundamentals

# ==================== PORTFOLIO MANAGEMENT ROUTES ====================

@api_router.get("/portfolios")
async def get_portfolios(user: User = Depends(get_current_user)):
    """Get all portfolios for the current user"""
    portfolios = await db.portfolios.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    
    # If no portfolios exist, create a default one
    if not portfolios:
        default_portfolio = Portfolio(
            user_id=user.user_id,
            name="Carteira Principal",
            description="Carteira padrão",
            is_default=True
        )
        doc = default_portfolio.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()
        doc["updated_at"] = doc["updated_at"].isoformat()
        await db.portfolios.insert_one(doc)
        portfolios = [doc]
        
        # Update existing stocks/dividends to use this portfolio
        await db.stocks.update_many(
            {"user_id": user.user_id, "$or": [{"portfolio_id": None}, {"portfolio_id": ""}, {"portfolio_id": {"$exists": False}}]},
            {"$set": {"portfolio_id": default_portfolio.portfolio_id}}
        )
        await db.dividends.update_many(
            {"user_id": user.user_id, "$or": [{"portfolio_id": None}, {"portfolio_id": ""}, {"portfolio_id": {"$exists": False}}]},
            {"$set": {"portfolio_id": default_portfolio.portfolio_id}}
        )
    
    # Add stats for each portfolio
    for portfolio in portfolios:
        portfolio_id = portfolio.get("portfolio_id")
        stocks_count = await db.stocks.count_documents({"user_id": user.user_id, "portfolio_id": portfolio_id})
        portfolio["stocks_count"] = stocks_count
    
    return portfolios

@api_router.post("/portfolios")
async def create_portfolio(portfolio_data: PortfolioCreate, user: User = Depends(get_current_user)):
    """Create a new portfolio"""
    portfolio = Portfolio(
        user_id=user.user_id,
        name=portfolio_data.name,
        description=portfolio_data.description,
        is_default=False
    )
    doc = portfolio.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    await db.portfolios.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/portfolios/{portfolio_id}")
async def update_portfolio(portfolio_id: str, portfolio_data: PortfolioUpdate, user: User = Depends(get_current_user)):
    """Update a portfolio"""
    update_data = {k: v for k, v in portfolio_data.model_dump().items() if v is not None}
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.portfolios.update_one(
            {"portfolio_id": portfolio_id, "user_id": user.user_id},
            {"$set": update_data}
        )
    portfolio = await db.portfolios.find_one({"portfolio_id": portfolio_id, "user_id": user.user_id}, {"_id": 0})
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    return portfolio

@api_router.delete("/portfolios/{portfolio_id}")
async def delete_portfolio(portfolio_id: str, user: User = Depends(get_current_user)):
    """Delete a portfolio and optionally move its stocks"""
    portfolio = await db.portfolios.find_one({"portfolio_id": portfolio_id, "user_id": user.user_id})
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    
    if portfolio.get("is_default"):
        raise HTTPException(status_code=400, detail="Cannot delete default portfolio")
    
    # Delete all stocks and dividends in this portfolio
    await db.stocks.delete_many({"portfolio_id": portfolio_id, "user_id": user.user_id})
    await db.dividends.delete_many({"portfolio_id": portfolio_id, "user_id": user.user_id})
    
    # Delete the portfolio
    await db.portfolios.delete_one({"portfolio_id": portfolio_id, "user_id": user.user_id})
    
    return {"message": "Portfolio deleted", "deleted_stocks": True}

@api_router.get("/portfolios/{portfolio_id}")
async def get_portfolio(portfolio_id: str, user: User = Depends(get_current_user)):
    """Get a specific portfolio with stats"""
    portfolio = await db.portfolios.find_one({"portfolio_id": portfolio_id, "user_id": user.user_id}, {"_id": 0})
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    
    # Add stats
    stocks = await db.stocks.find({"portfolio_id": portfolio_id, "user_id": user.user_id}, {"_id": 0}).to_list(1000)
    dividends = await db.dividends.find({"portfolio_id": portfolio_id, "user_id": user.user_id}, {"_id": 0}).to_list(10000)
    
    portfolio["stocks_count"] = len(stocks)
    portfolio["total_invested"] = sum(s["quantity"] * s["average_price"] for s in stocks)
    portfolio["total_current"] = sum(s["quantity"] * (s.get("current_price") or s["average_price"]) for s in stocks)
    portfolio["total_dividends"] = sum(d["amount"] for d in dividends)
    
    return portfolio

# ==================== PORTFOLIO STOCKS ROUTES ====================

@api_router.get("/portfolio/stocks")
async def get_stocks(user: User = Depends(get_current_user), portfolio_id: Optional[str] = None):
    """Get stocks, optionally filtered by portfolio"""
    query = {"user_id": user.user_id}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    else:
        # If no portfolio_id specified, get all stocks (for backwards compatibility)
        pass
    stocks = await db.stocks.find(query, {"_id": 0}).to_list(1000)
    return stocks

@api_router.post("/portfolio/stocks")
async def add_stock(stock_data: StockCreate, user: User = Depends(get_current_user)):
    # Get or create default portfolio if no portfolio_id provided
    portfolio_id = stock_data.portfolio_id
    if not portfolio_id:
        # Get default portfolio
        default_portfolio = await db.portfolios.find_one({"user_id": user.user_id, "is_default": True})
        if not default_portfolio:
            # Create default portfolio
            default_portfolio = Portfolio(
                user_id=user.user_id,
                name="Carteira Principal",
                description="Carteira padrão",
                is_default=True
            )
            doc = default_portfolio.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            doc["updated_at"] = doc["updated_at"].isoformat()
            await db.portfolios.insert_one(doc)
            portfolio_id = default_portfolio.portfolio_id
        else:
            portfolio_id = default_portfolio.get("portfolio_id")
    
    stock = Stock(
        user_id=user.user_id,
        portfolio_id=portfolio_id,
        ticker=stock_data.ticker.upper(),
        name=stock_data.name,
        quantity=stock_data.quantity,
        average_price=stock_data.average_price,
        purchase_date=stock_data.purchase_date,
        sector=stock_data.sector,
        current_price=stock_data.current_price,
        dividend_yield=stock_data.dividend_yield,
        ceiling_price=stock_data.ceiling_price
    )
    doc = stock.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    await db.stocks.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/portfolio/stocks/{stock_id}")
async def update_stock(stock_id: str, stock_data: StockUpdate, user: User = Depends(get_current_user)):
    update_fields = {k: v for k, v in stock_data.model_dump().items() if v is not None}
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.stocks.update_one(
            {"stock_id": stock_id, "user_id": user.user_id},
            {"$set": update_fields}
        )
    stock = await db.stocks.find_one({"stock_id": stock_id, "user_id": user.user_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stock not found")
    return stock

@api_router.delete("/portfolio/stocks/all")
async def delete_all_stocks(user: User = Depends(get_current_user)):
    """Delete all stocks for the current user"""
    result = await db.stocks.delete_many({"user_id": user.user_id})
    # Also delete related dividends
    await db.dividends.delete_many({"user_id": user.user_id})
    return {"message": f"{result.deleted_count} ações excluídas", "deleted": result.deleted_count}

@api_router.delete("/portfolio/stocks/{stock_id}")
async def delete_stock(stock_id: str, user: User = Depends(get_current_user)):
    result = await db.stocks.delete_one({"stock_id": stock_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stock not found")
    return {"message": "Stock deleted"}

@api_router.get("/portfolio/summary")
async def get_portfolio_summary(user: User = Depends(get_current_user)):
    stocks = await db.stocks.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    
    total_invested = sum(s["quantity"] * s["average_price"] for s in stocks)
    total_current = sum(s["quantity"] * (s.get("current_price") or s["average_price"]) for s in stocks)
    total_gain = total_current - total_invested
    gain_percent = (total_gain / total_invested * 100) if total_invested > 0 else 0
    
    # Get dividends and filter only those already paid
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    dividends = await db.dividends.find({"user_id": user.user_id}, {"_id": 0}).to_list(10000)
    
    # Only sum dividends where payment_date <= today
    total_dividends_received = sum(
        d["amount"] for d in dividends 
        if d.get("payment_date", "")[:10] <= today
    )
    total_dividends_pending = sum(
        d["amount"] for d in dividends 
        if d.get("payment_date", "")[:10] > today
    )
    
    return {
        "total_invested": round(total_invested, 2),
        "total_current": round(total_current, 2),
        "total_gain": round(total_gain, 2),
        "gain_percent": round(gain_percent, 2),
        "total_dividends": round(total_dividends_received, 2),  # Only received dividends
        "total_dividends_pending": round(total_dividends_pending, 2),  # Pending dividends
        "stocks_count": len(stocks)
    }

@api_router.post("/portfolio/refresh-prices")
async def refresh_portfolio_prices(user: User = Depends(get_current_user)):
    """Refresh all stock prices - uses Yahoo Finance as primary source"""
    stocks = await db.stocks.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    updated = 0
    alerts_created = 0
    errors = []
    
    # Get unique tickers to avoid redundant API calls
    unique_tickers = list(set(stock["ticker"] for stock in stocks))
    ticker_prices = {}  # Cache prices by ticker
    
    # Fetch price for each unique ticker
    for ticker in unique_tickers:
        try:
            new_price = None
            source = None
            
            # Try Yahoo Finance first (most reliable for B3 stocks)
            yahoo_data = await fetch_yahoo_finance_quote(ticker)
            if yahoo_data and yahoo_data["price"] > 0:
                new_price = yahoo_data["price"]
                source = "yahoo_finance"
            else:
                # Fallback to TradingView
                tv_data = fetch_tradingview_quote(ticker)
                if tv_data and tv_data["price"] > 0:
                    new_price = tv_data["price"]
                    source = "tradingview"
                else:
                    # Last resort: Alpha Vantage
                    await asyncio.sleep(0.3)  # Delay to avoid rate limit
                    av_data = await fetch_alpha_vantage_quote(ticker)
                    if av_data and av_data["price"] > 0:
                        new_price = av_data["price"]
                        source = "alpha_vantage"
            
            if new_price:
                ticker_prices[ticker] = {"price": new_price, "source": source}
                logger.info(f"Fetched {ticker}: R${new_price:.2f} (source: {source})")
            else:
                errors.append(ticker)
                logger.warning(f"Could not fetch price for {ticker}")
            
            # Small delay between requests to be respectful
            await asyncio.sleep(0.1)
                
        except Exception as e:
            logger.error(f"Error fetching price for {ticker}: {e}")
            errors.append(ticker)
    
    # Update all stock records with fetched prices
    for stock in stocks:
        ticker = stock["ticker"]
        if ticker in ticker_prices:
            new_price = ticker_prices[ticker]["price"]
            
            # Check for ceiling price alerts (only once per ticker)
            if stock.get("ceiling_price") and new_price >= stock["ceiling_price"]:
                today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                existing_alert = await db.alerts.find_one({
                    "user_id": user.user_id,
                    "ticker": ticker,
                    "alert_type": "ceiling_reached",
                    "created_at": {"$regex": f"^{today}"}
                })
                
                if not existing_alert:
                    alert = Alert(
                        user_id=user.user_id,
                        stock_id=stock["stock_id"],
                        ticker=ticker,
                        alert_type="ceiling_reached",
                        message=f"{ticker} atingiu o preço teto! Atual: R${new_price:.2f}, Teto: R${stock['ceiling_price']:.2f}"
                    )
                    alert_doc = alert.model_dump()
                    alert_doc["created_at"] = alert_doc["created_at"].isoformat()
                    await db.alerts.insert_one(alert_doc)
                    alerts_created += 1
            
            # Update this stock record
            await db.stocks.update_one(
                {"stock_id": stock["stock_id"]},
                {"$set": {"current_price": new_price, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            updated += 1
    
    # Save portfolio snapshot
    await save_portfolio_snapshot(user.user_id)
    
    result = {
        "updated": updated, 
        "total": len(stocks),
        "unique_tickers": len(unique_tickers),
        "alerts_created": alerts_created,
        "source": "tradingview/alpha_vantage"
    }
    if errors:
        result["errors"] = list(set(errors))  # Remove duplicates
    
    return result

# ==================== PORTFOLIO HISTORY ====================

async def save_portfolio_snapshot(user_id: str):
    """Save daily portfolio snapshot for evolution chart"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Check if snapshot already exists for today
    existing = await db.portfolio_snapshots.find_one(
        {"user_id": user_id, "date": today}, {"_id": 0}
    )
    if existing:
        return existing
    
    stocks = await db.stocks.find({"user_id": user_id}, {"_id": 0}).to_list(1000)
    dividends = await db.dividends.find({"user_id": user_id}, {"_id": 0}).to_list(10000)
    
    total_invested = sum(s["quantity"] * s["average_price"] for s in stocks)
    total_current = sum(s["quantity"] * (s.get("current_price") or s["average_price"]) for s in stocks)
    total_dividends = sum(d["amount"] for d in dividends)
    
    snapshot = PortfolioSnapshot(
        user_id=user_id,
        date=today,
        total_invested=round(total_invested, 2),
        total_current=round(total_current, 2),
        total_dividends=round(total_dividends, 2),
        stocks_count=len(stocks)
    )
    
    doc = snapshot.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.portfolio_snapshots.insert_one(doc)
    
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/portfolio/history")
async def get_portfolio_history(user: User = Depends(get_current_user), days: int = 30):
    """Get portfolio evolution history"""
    snapshots = await db.portfolio_snapshots.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).sort("date", -1).limit(days).to_list(days)
    
    return list(reversed(snapshots))

@api_router.get("/portfolio/evolution")
async def get_portfolio_evolution(user: User = Depends(get_current_user), period: str = "1m", portfolio_id: Optional[str] = None):
    """
    Get portfolio evolution based on purchase dates.
    Period options: 1w (week), 1m (month), 12m (year), 5y (5 years), max (all time)
    """
    query = {"user_id": user.user_id}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    
    stocks = await db.stocks.find(query, {"_id": 0}).to_list(1000)
    dividends = await db.dividends.find(query, {"_id": 0}).to_list(10000)
    
    if not stocks:
        return []
    
    # Calculate date range based on period
    today = datetime.now(timezone.utc).date()
    
    period_days = {
        "1w": 7,
        "1m": 30,
        "12m": 365,
        "5y": 1825,
        "max": 3650  # 10 years max
    }
    
    days_back = period_days.get(period, 30)
    start_date = today - timedelta(days=days_back)
    
    # Find the earliest purchase date
    earliest_purchase = None
    for stock in stocks:
        if stock.get("purchase_date"):
            try:
                purchase_dt = datetime.strptime(stock["purchase_date"], "%Y-%m-%d").date()
                if earliest_purchase is None or purchase_dt < earliest_purchase:
                    earliest_purchase = purchase_dt
            except:
                pass
    
    # If we have purchases, start from the earliest one (within our period)
    if earliest_purchase and earliest_purchase > start_date:
        start_date = earliest_purchase
    
    # Create dividend lookup by payment_date (only include paid dividends)
    dividend_by_date = {}
    for div in dividends:
        payment_date = div.get("payment_date", "")[:10]
        if payment_date:
            # Only include dividends with payment_date in the past (already paid)
            try:
                payment_dt = datetime.strptime(payment_date, "%Y-%m-%d").date()
                if payment_dt <= today:  # Only include already paid dividends
                    dividend_by_date[payment_date] = dividend_by_date.get(payment_date, 0) + div["amount"]
            except:
                pass
    
    # Generate evolution data points
    evolution = []
    current_date = start_date
    cumulative_dividends = 0
    
    # Pre-calculate dividends received before start_date (only paid dividends)
    for div in dividends:
        payment_date = div.get("payment_date", "")[:10]
        if payment_date:
            try:
                div_date = datetime.strptime(payment_date, "%Y-%m-%d").date()
                # Only include if already paid AND before start_date
                if div_date < start_date and div_date <= today:
                    cumulative_dividends += div["amount"]
            except:
                pass
    
    while current_date <= today:
        date_str = current_date.strftime("%Y-%m-%d")
        
        # Calculate portfolio value on this date
        # Only include stocks that were purchased on or before this date
        total_invested = 0
        total_current = 0
        
        for stock in stocks:
            purchase_date = stock.get("purchase_date")
            if purchase_date:
                try:
                    purchase_dt = datetime.strptime(purchase_date, "%Y-%m-%d").date()
                    if purchase_dt <= current_date:
                        # Stock was owned on this date
                        invested = stock["quantity"] * stock["average_price"]
                        current_value = stock["quantity"] * (stock.get("current_price") or stock["average_price"])
                        total_invested += invested
                        total_current += current_value
                except:
                    # If date parsing fails, include the stock
                    total_invested += stock["quantity"] * stock["average_price"]
                    total_current += stock["quantity"] * (stock.get("current_price") or stock["average_price"])
            else:
                # No purchase date, include the stock
                total_invested += stock["quantity"] * stock["average_price"]
                total_current += stock["quantity"] * (stock.get("current_price") or stock["average_price"])
        
        # Add dividends received on this date
        if date_str in dividend_by_date:
            cumulative_dividends += dividend_by_date[date_str]
        
        # Only add data points where we have investments
        if total_invested > 0:
            evolution.append({
                "date": date_str,
                "invested": round(total_invested, 2),
                "current": round(total_current, 2),
                "dividends": round(cumulative_dividends, 2),
                "total": round(total_current + cumulative_dividends, 2),
                "gain": round(total_current - total_invested, 2),
                "gain_percent": round(((total_current - total_invested) / total_invested) * 100, 2) if total_invested > 0 else 0
            })
        
        # Move to next date (adjust granularity based on period)
        if period == "1w":
            current_date += timedelta(days=1)
        elif period == "1m":
            current_date += timedelta(days=1)
        elif period == "12m":
            current_date += timedelta(days=7)  # Weekly for 12 months
        elif period == "5y":
            current_date += timedelta(days=30)  # Monthly for 5 years
        else:  # max
            current_date += timedelta(days=30)  # Monthly
    
    return evolution

@api_router.post("/portfolio/snapshot")
async def create_portfolio_snapshot(user: User = Depends(get_current_user)):
    """Manually create a portfolio snapshot"""
    return await save_portfolio_snapshot(user.user_id)

# ==================== IMPORT CEI/B3 ====================

def parse_cei_csv(content: str) -> List[dict]:
    """Parse CEI/B3 CSV export file - groups by ticker + purchase_date"""
    stocks = []
    stocks_dict = {}  # Key: (ticker, purchase_date)
    
    # Try different delimiters - TAB is common in CEI exports
    for delimiter in ['\t', ';', ',']:
        try:
            reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
            headers = reader.fieldnames
            
            if not headers:
                continue
            
            logger.info(f"CEI parser trying delimiter '{delimiter}', headers: {headers}")
            
            # Check for CEI format headers
            produto_col = None
            qtd_col = None
            preco_col = None
            data_col = None
            
            for h in headers:
                h_lower = h.lower().strip()
                # Handle encoding issues
                h_normalized = h_lower.replace('ã', 'a').replace('ç', 'c').replace('í', 'i').replace('ú', 'u').replace('á', 'a').replace('é', 'e').replace('ó', 'o')
                
                # Look for ticker/code column - check for "Código de Negociação" first
                if produto_col is None:
                    if 'codigo' in h_normalized and 'negociacao' in h_normalized:
                        produto_col = h
                    elif 'codigo' in h_normalized or 'código' in h_normalized:
                        produto_col = h
                    elif 'produto' in h_normalized or 'ativo' in h_normalized:
                        produto_col = h
                    elif 'ticker' in h_normalized or 'papel' in h_normalized:
                        produto_col = h
                
                # Look for quantity column
                if qtd_col is None:
                    if 'quantidade' in h_normalized or 'qtd' in h_normalized:
                        qtd_col = h
                
                # Look for price column
                if preco_col is None:
                    if 'preco' in h_normalized or 'preço' in h_normalized:
                        preco_col = h
                    elif 'unitario' in h_normalized or 'unitário' in h_normalized:
                        preco_col = h
                    elif 'medio' in h_normalized or 'médio' in h_normalized:
                        preco_col = h
                
                # Look for date columns - "Data do Negócio"
                if data_col is None:
                    if 'data' in h_normalized:
                        if 'negocio' in h_normalized or 'negociacao' in h_normalized:
                            data_col = h
                        elif 'compra' in h_normalized or 'aquisicao' in h_normalized:
                            data_col = h
                        elif h_normalized in ['data', 'date', 'dt', 'data do negocio']:
                            data_col = h
            
            logger.info(f"CEI columns found - Produto: {produto_col}, Qtd: {qtd_col}, Preco: {preco_col}, Data: {data_col}")
            
            if not produto_col:
                continue
            
            for row in reader:
                try:
                    produto = row.get(produto_col, '').strip()
                    if not produto:
                        continue
                    
                    # Extract ticker from "FIQE3 - UNIFIQUE TELECOMUNICAÇÕES S.A." format
                    ticker = produto.split(' - ')[0].split(' ')[0].strip().upper()
                    ticker = re.sub(r'[^A-Z0-9]', '', ticker)
                    
                    # Remove trailing "F" (fracionário) from ticker
                    if ticker.endswith('F') and len(ticker) > 4:
                        ticker = ticker[:-1]
                    
                    if not ticker or len(ticker) < 4:
                        continue
                    
                    # Get name from produto
                    name_parts = produto.split(' - ')
                    name = name_parts[1].strip() if len(name_parts) > 1 else ticker
                    
                    # Parse quantity
                    quantity = 0
                    if qtd_col and row.get(qtd_col):
                        qty_str = str(row.get(qtd_col, '0')).strip()
                        qty_str = re.sub(r'[^\d.,\-]', '', qty_str)
                        if qty_str:
                            if ',' in qty_str and '.' in qty_str:
                                qty_str = qty_str.replace('.', '').replace(',', '.')
                            elif ',' in qty_str:
                                qty_str = qty_str.replace(',', '.')
                            quantity = float(qty_str)
                    
                    # Parse price
                    avg_price = 0
                    if preco_col and row.get(preco_col):
                        price_str = str(row.get(preco_col, '0')).strip()
                        # Remove R$, spaces, quotes
                        price_str = re.sub(r'[R$\s"\']', '', price_str)
                        if price_str:
                            if ',' in price_str and '.' in price_str:
                                price_str = price_str.replace('.', '').replace(',', '.')
                            elif ',' in price_str:
                                price_str = price_str.replace(',', '.')
                            avg_price = float(price_str)
                    
                    # Parse purchase date
                    purchase_date = None
                    if data_col and row.get(data_col):
                        date_str = str(row.get(data_col, '')).strip()
                        if date_str:
                            # Try different date formats
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y']:
                                try:
                                    parsed_date = datetime.strptime(date_str, fmt)
                                    purchase_date = parsed_date.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    continue
                    
                    # Group by ticker + purchase_date
                    # Same ticker on same date = aggregate
                    # Same ticker on different dates = separate records
                    key = (ticker, purchase_date)
                    
                    if key in stocks_dict:
                        old_qty = stocks_dict[key]['quantity']
                        old_price = stocks_dict[key]['average_price']
                        new_qty = old_qty + quantity
                        # Calculate weighted average price
                        if new_qty > 0 and avg_price > 0:
                            stocks_dict[key]['average_price'] = ((old_qty * old_price) + (quantity * avg_price)) / new_qty
                        stocks_dict[key]['quantity'] = new_qty
                    else:
                        stocks_dict[key] = {
                            "ticker": ticker,
                            "name": name,
                            "quantity": quantity,
                            "average_price": avg_price,
                            "purchase_date": purchase_date
                        }
                        
                except Exception as e:
                    logger.error(f"Error parsing CEI row: {e}")
                    continue
            
            if stocks_dict:
                stocks = [s for s in stocks_dict.values() if s['quantity'] > 0]
                logger.info(f"CEI parser found {len(stocks)} stock entries (grouped by ticker+date)")
                break
                
        except Exception as e:
            logger.error(f"CEI parser error: {e}")
            continue
    
    return stocks

def parse_xlsx(file_bytes: bytes) -> List[dict]:
    """Parse Excel XLSX file - groups by ticker + purchase_date"""
    stocks = []
    stocks_dict = {}  # Key: (ticker, purchase_date)
    
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=False, data_only=True)
        ws = wb.active
        
        # Find header row (might not be the first row)
        headers = []
        header_row = 0
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
            # Check if this row looks like a header
            row_values = [str(cell or '').strip().lower() for cell in row if cell]
            row_str = ' '.join(row_values)
            
            # Look for keywords that indicate this is a header row
            if any(kw in row_str for kw in ['produto', 'ticker', 'codigo', 'ativo', 'quantidade', 'preco', 'preço']):
                headers = [str(cell or '').strip() for cell in row]
                header_row = row_idx
                logger.info(f"XLSX found header at row {row_idx}: {headers}")
                break
        
        if not headers:
            # If no header found, use first row
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(cell or '').strip() for cell in row]
            logger.info(f"XLSX using first row as headers: {headers}")
        
        logger.info(f"XLSX headers: {headers}")
        
        # Map columns - be very specific to avoid false matches
        def find_col_idx(key):
            # First pass: look for exact/specific matches
            exact_matches = {
                'ticker': ['código de negociação', 'codigo de negociacao', 'ticker', 'código', 'codigo', 'papel'],
                'name': ['nome', 'empresa', 'name'],
                'quantity': ['quantidade', 'qtd'],
                'average_price': ['preço', 'preco', 'preço unitário', 'preco unitario'],
                'purchase_date': ['data do negócio', 'data do negocio', 'data'],
                'sector': ['setor', 'sector']
            }
            
            for idx, h in enumerate(headers):
                if not h:
                    continue
                h_lower = h.lower().strip()
                h_normalized = h_lower.replace('ã', 'a').replace('ç', 'c').replace('í', 'i').replace('é', 'e').replace('ú', 'u').replace('á', 'a').replace('ó', 'o')
                
                for keyword in exact_matches.get(key, []):
                    keyword_normalized = keyword.replace('ã', 'a').replace('ç', 'c').replace('í', 'i').replace('é', 'e').replace('ú', 'u').replace('á', 'a').replace('ó', 'o')
                    # Check for exact match or if header equals keyword
                    if h_normalized == keyword_normalized:
                        return idx
                    # Check if header contains the full keyword (not partial)
                    if keyword_normalized in h_normalized and len(keyword_normalized) >= 4:
                        return idx
            return None
        
        ticker_idx = find_col_idx('ticker')
        name_idx = find_col_idx('name')
        qty_idx = find_col_idx('quantity')
        price_idx = find_col_idx('average_price')
        date_idx = find_col_idx('purchase_date')
        sector_idx = find_col_idx('sector')
        
        logger.info(f"XLSX column indices - ticker: {ticker_idx}, qty: {qty_idx}, price: {price_idx}, date: {date_idx}")
        
        if ticker_idx is None:
            logger.error("No ticker column found in XLSX")
            wb.close()
            return []
        
        # Process rows (skip header row)
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 2, values_only=True)):
            try:
                if not row or len(row) <= ticker_idx:
                    continue
                
                produto = str(row[ticker_idx] or '').strip()
                if not produto:
                    continue
                
                # Extract ticker from "FIQE3 - UNIFIQUE..." format
                ticker = produto.split(' - ')[0].split(' ')[0].strip().upper()
                ticker = re.sub(r'[^A-Z0-9]', '', ticker)
                
                # Remove trailing "F" (fracionário) from ticker
                if ticker.endswith('F') and len(ticker) > 4:
                    ticker = ticker[:-1]
                
                if not ticker or len(ticker) < 4:
                    continue
                
                # Get name
                name = ticker
                if name_idx is not None and len(row) > name_idx and row[name_idx]:
                    name = str(row[name_idx]).strip()
                elif ' - ' in produto:
                    name = produto.split(' - ')[1].strip()
                
                # Parse quantity
                quantity = 0
                if qty_idx is not None and len(row) > qty_idx and row[qty_idx] is not None:
                    qty_val = row[qty_idx]
                    if isinstance(qty_val, (int, float)):
                        quantity = float(qty_val)
                    else:
                        qty_str = re.sub(r'[^\d.,\-]', '', str(qty_val))
                        if qty_str:
                            if ',' in qty_str and '.' in qty_str:
                                qty_str = qty_str.replace('.', '').replace(',', '.')
                            elif ',' in qty_str:
                                qty_str = qty_str.replace(',', '.')
                            quantity = float(qty_str) if qty_str else 0
                
                # Parse price
                avg_price = 0
                if price_idx is not None and len(row) > price_idx and row[price_idx] is not None:
                    price_val = row[price_idx]
                    if isinstance(price_val, (int, float)):
                        avg_price = float(price_val)
                    else:
                        price_str = re.sub(r'[R$\s]', '', str(price_val))
                        if price_str:
                            if ',' in price_str and '.' in price_str:
                                price_str = price_str.replace('.', '').replace(',', '.')
                            elif ',' in price_str:
                                price_str = price_str.replace(',', '.')
                            avg_price = float(price_str) if price_str else 0
                
                # Parse purchase date
                purchase_date = None
                if date_idx is not None and len(row) > date_idx and row[date_idx] is not None:
                    date_val = row[date_idx]
                    if isinstance(date_val, datetime):
                        purchase_date = date_val.strftime('%Y-%m-%d')
                    elif date_val:
                        date_str = str(date_val).strip()
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y']:
                            try:
                                parsed_date = datetime.strptime(date_str, fmt)
                                purchase_date = parsed_date.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                
                # Group by ticker + purchase_date
                # Same ticker on same date = aggregate
                # Same ticker on different dates = separate records
                key = (ticker, purchase_date)
                
                if key in stocks_dict:
                    old_qty = stocks_dict[key]['quantity']
                    old_price = stocks_dict[key]['average_price']
                    new_qty = old_qty + quantity
                    if new_qty > 0 and avg_price > 0:
                        stocks_dict[key]['average_price'] = ((old_qty * old_price) + (quantity * avg_price)) / new_qty
                    stocks_dict[key]['quantity'] = new_qty
                else:
                    stocks_dict[key] = {
                        "ticker": ticker,
                        "name": name,
                        "quantity": quantity,
                        "average_price": avg_price,
                        "purchase_date": purchase_date,
                        "sector": None
                    }
                    
            except Exception as e:
                logger.error(f"Error parsing XLSX row {row_idx}: {e}")
                continue
        
        wb.close()
        stocks = [s for s in stocks_dict.values() if s['quantity'] > 0]
        logger.info(f"XLSX parser found {len(stocks)} stock entries (grouped by ticker+date)")
        
    except Exception as e:
        logger.error(f"XLSX parser error: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    return stocks

def parse_generic_csv(content: str) -> List[dict]:
    """Parse generic CSV format - groups by ticker + purchase_date"""
    stocks = []
    stocks_dict = {}  # Key: (ticker, purchase_date)
    
    # Try different delimiters
    for delimiter in [',', ';', '\t', '|']:
        try:
            reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
            headers = reader.fieldnames
            
            if not headers:
                continue
            
            logger.info(f"CSV headers found with delimiter '{delimiter}': {headers}")
            
            # Map common header variations (case insensitive)
            header_map = {
                'ticker': ['ticker', 'codigo', 'código', 'symbol', 'ativo', 'papel', 'acao', 'ação', 'code', 'stock'],
                'name': ['name', 'nome', 'empresa', 'description', 'descricao', 'descrição', 'produto', 'ativo'],
                'quantity': ['quantity', 'quantidade', 'qtd', 'qtde', 'shares', 'qty', 'quant'],
                'average_price': ['average_price', 'preco_medio', 'preço_médio', 'avg_price', 'cost', 'preco', 'preço', 'pm', 'custo', 'valor'],
                'purchase_date': ['purchase_date', 'data_compra', 'date', 'data', 'dt_compra'],
                'sector': ['sector', 'setor', 'industry', 'segmento']
            }
            
            def find_header(key):
                for h in headers:
                    h_lower = h.lower().strip().replace(' ', '_').replace('-', '_')
                    if h_lower in header_map[key]:
                        return h
                    # Also check if header contains any of the keywords
                    for keyword in header_map[key]:
                        if keyword in h_lower:
                            return h
                return None
            
            ticker_col = find_header('ticker')
            name_col = find_header('name')
            qty_col = find_header('quantity')
            price_col = find_header('average_price')
            date_col = find_header('purchase_date')
            sector_col = find_header('sector')
            
            logger.info(f"CSV columns found - Ticker: {ticker_col}, Qty: {qty_col}, Price: {price_col}, Date: {date_col}")
            
            if not ticker_col:
                continue
            
            for row in reader:
                try:
                    ticker = row.get(ticker_col, '').strip().upper()
                    # Remove common suffixes and clean ticker
                    ticker = re.sub(r'[^A-Z0-9]', '', ticker)
                    
                    # Remove trailing "F" (fracionário) from ticker
                    if ticker.endswith('F') and len(ticker) > 4:
                        ticker = ticker[:-1]
                    
                    if not ticker or len(ticker) < 4:
                        continue
                    
                    name = row.get(name_col, ticker) if name_col else ticker
                    
                    # Parse quantity
                    quantity = 0
                    if qty_col and row.get(qty_col):
                        qty_str = str(row.get(qty_col, '0')).strip()
                        # Handle Brazilian number format (1.000,50 -> 1000.50)
                        if ',' in qty_str and '.' in qty_str:
                            qty_str = qty_str.replace('.', '').replace(',', '.')
                        elif ',' in qty_str:
                            qty_str = qty_str.replace(',', '.')
                        quantity = float(qty_str) if qty_str else 0
                    
                    # Parse price
                    avg_price = 0
                    if price_col and row.get(price_col):
                        price_str = str(row.get(price_col, '0')).strip()
                        # Remove currency symbols
                        price_str = re.sub(r'[R$\s]', '', price_str)
                        # Handle Brazilian number format
                        if ',' in price_str and '.' in price_str:
                            price_str = price_str.replace('.', '').replace(',', '.')
                        elif ',' in price_str:
                            price_str = price_str.replace(',', '.')
                        avg_price = float(price_str) if price_str else 0
                    
                    # Parse purchase date
                    purchase_date = None
                    if date_col and row.get(date_col):
                        date_str = str(row.get(date_col, '')).strip()
                        if date_str:
                            # Try different date formats
                            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y']:
                                try:
                                    parsed_date = datetime.strptime(date_str, fmt)
                                    purchase_date = parsed_date.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    continue
                    
                    sector = row.get(sector_col) if sector_col else None
                    
                    # Group by ticker + purchase_date
                    # Same ticker on same date = aggregate
                    # Same ticker on different dates = separate records
                    key = (ticker, purchase_date)
                    
                    if key in stocks_dict:
                        old_qty = stocks_dict[key]['quantity']
                        old_price = stocks_dict[key]['average_price']
                        new_qty = old_qty + quantity
                        # Calculate weighted average price
                        if new_qty > 0 and avg_price > 0:
                            stocks_dict[key]['average_price'] = ((old_qty * old_price) + (quantity * avg_price)) / new_qty
                        stocks_dict[key]['quantity'] = new_qty
                    else:
                        stocks_dict[key] = {
                            "ticker": ticker,
                            "name": name,
                            "quantity": quantity,
                            "average_price": avg_price,
                            "purchase_date": purchase_date,
                            "sector": sector
                        }
                        
                except Exception as e:
                    logger.error(f"Error parsing row: {e}")
                    continue
            
            if stocks_dict:
                stocks = [s for s in stocks_dict.values() if s['quantity'] > 0]
                logger.info(f"CSV parser found {len(stocks)} stock entries (grouped by ticker+date)")
                break
        except Exception as e:
            logger.error(f"Error with delimiter '{delimiter}': {e}")
            continue
    
    return stocks

@api_router.post("/portfolio/import")
async def import_file(file: UploadFile = File(...), portfolio_id: Optional[str] = Form(None), user: User = Depends(get_current_user)):
    """Import stocks from CSV or XLSX file"""
    content = await file.read()
    filename = file.filename or ""
    
    logger.info(f"Importing file: {filename}, size: {len(content)} bytes, portfolio_id: {portfolio_id}")
    
    # Get or create default portfolio if no portfolio_id provided
    if not portfolio_id:
        default_portfolio = await db.portfolios.find_one({"user_id": user.user_id, "is_default": True})
        if not default_portfolio:
            # Create default portfolio
            new_portfolio = Portfolio(user_id=user.user_id, name="Minha Carteira", is_default=True)
            doc = new_portfolio.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            doc["updated_at"] = doc["updated_at"].isoformat()
            await db.portfolios.insert_one(doc)
            portfolio_id = new_portfolio.portfolio_id
        else:
            portfolio_id = default_portfolio.get("portfolio_id")
    
    stocks = []
    
    # Check if it's an Excel file
    if filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls'):
        stocks = parse_xlsx(content)
    else:
        # Try to parse as CSV
        content_str = None
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                content_str = content.decode(encoding)
                logger.info(f"Successfully decoded with {encoding}")
                break
            except:
                continue
        
        if not content_str:
            raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo. Codificação não suportada.")
        
        # Log first few lines for debugging
        lines = content_str.split('\n')[:3]
        logger.info(f"First lines of file: {lines}")
        
        # Try CEI format first
        stocks = parse_cei_csv(content_str)
        logger.info(f"CEI parser found {len(stocks)} stocks")
        
        # If no stocks found, try generic format
        if not stocks:
            stocks = parse_generic_csv(content_str)
            logger.info(f"Generic parser found {len(stocks)} stocks")
    
    if not stocks:
        raise HTTPException(
            status_code=400, 
            detail="Não foi possível ler o arquivo. Verifique se contém uma coluna 'ticker' ou 'produto'. Formatos aceitos: CSV, XLSX"
        )
    
    imported = 0
    updated = 0
    
    for stock_data in stocks:
        ticker = stock_data["ticker"]
        purchase_date = stock_data.get("purchase_date")
        
        # Check if stock already exists with same ticker AND purchase_date
        # This allows multiple entries for the same stock bought on different dates
        query = {"user_id": user.user_id, "ticker": ticker}
        if purchase_date:
            query["purchase_date"] = purchase_date
        
        existing = await db.stocks.find_one(query, {"_id": 0})
        
        # Get additional info
        stock_info = BRAZILIAN_STOCKS.get(ticker, {})
        
        if existing:
            # Update existing stock (same ticker + same date)
            await db.stocks.update_one(
                {"stock_id": existing["stock_id"]},
                {"$set": {
                    "quantity": stock_data["quantity"],
                    "average_price": stock_data["average_price"],
                    "purchase_date": purchase_date,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            updated += 1
        else:
            # Create new stock entry
            new_stock = Stock(
                user_id=user.user_id,
                ticker=ticker,
                name=stock_data.get("name") or stock_info.get("name", ticker),
                quantity=stock_data["quantity"],
                average_price=stock_data["average_price"],
                purchase_date=purchase_date,
                sector=stock_data.get("sector") or stock_info.get("sector"),
                current_price=stock_info.get("current_price"),
                dividend_yield=stock_info.get("dividend_yield")
            )
            doc = new_stock.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            doc["updated_at"] = doc["updated_at"].isoformat()
            await db.stocks.insert_one(doc)
            imported += 1
    
    return {
        "imported": imported,
        "updated": updated,
        "total": len(stocks),
        "message": f"Importação concluída: {imported} novas ações, {updated} atualizadas"
    }

@api_router.get("/portfolio/export/csv")
async def export_csv(user: User = Depends(get_current_user)):
    """Export portfolio to CSV"""
    stocks = await db.stocks.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ticker', 'name', 'quantity', 'average_price', 'current_price', 'purchase_date', 'sector', 'ceiling_price'])
    
    for stock in stocks:
        writer.writerow([
            stock['ticker'],
            stock['name'],
            stock['quantity'],
            stock['average_price'],
            stock.get('current_price', ''),
            stock.get('purchase_date', ''),
            stock.get('sector', ''),
            stock.get('ceiling_price', '')
        ])
    
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=portfolio.csv"}
    )

# ==================== ALERTS ====================

@api_router.get("/alerts")
async def get_alerts(user: User = Depends(get_current_user), unread_only: bool = False):
    """Get user alerts"""
    query = {"user_id": user.user_id}
    if unread_only:
        query["is_read"] = False
    
    alerts = await db.alerts.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return alerts

@api_router.put("/alerts/{alert_id}/read")
async def mark_alert_read(alert_id: str, user: User = Depends(get_current_user)):
    """Mark alert as read"""
    result = await db.alerts.update_one(
        {"alert_id": alert_id, "user_id": user.user_id},
        {"$set": {"is_read": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Alert not found")
    return {"message": "Alert marked as read"}

@api_router.put("/alerts/read-all")
async def mark_all_alerts_read(user: User = Depends(get_current_user)):
    """Mark all alerts as read"""
    result = await db.alerts.update_many(
        {"user_id": user.user_id, "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": f"{result.modified_count} alerts marked as read"}

@api_router.delete("/alerts/{alert_id}")
async def delete_alert(alert_id: str, user: User = Depends(get_current_user)):
    """Delete an alert"""
    result = await db.alerts.delete_one({"alert_id": alert_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Alert not found")
    return {"message": "Alert deleted"}

@api_router.get("/alerts/count")
async def get_unread_alerts_count(user: User = Depends(get_current_user)):
    """Get count of unread alerts"""
    count = await db.alerts.count_documents({"user_id": user.user_id, "is_read": False})
    return {"count": count}

# ==================== DIVIDENDS ROUTES ====================

@api_router.get("/dividends")
async def get_dividends(user: User = Depends(get_current_user), portfolio_id: Optional[str] = None):
    query = {"user_id": user.user_id}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    dividends = await db.dividends.find(query, {"_id": 0}).to_list(10000)
    return dividends

@api_router.post("/dividends")
async def add_dividend(dividend_data: DividendCreate, user: User = Depends(get_current_user)):
    # Get portfolio_id from the stock or use provided one
    portfolio_id = dividend_data.portfolio_id
    if not portfolio_id and dividend_data.stock_id:
        stock = await db.stocks.find_one({"stock_id": dividend_data.stock_id, "user_id": user.user_id})
        if stock:
            portfolio_id = stock.get("portfolio_id")
    
    dividend = Dividend(
        user_id=user.user_id,
        stock_id=dividend_data.stock_id,
        ticker=dividend_data.ticker.upper(),
        portfolio_id=portfolio_id,
        amount=dividend_data.amount,
        payment_date=dividend_data.payment_date,
        type=dividend_data.type
    )
    doc = dividend.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.dividends.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/dividends/summary")
async def get_dividends_summary(user: User = Depends(get_current_user), portfolio_id: Optional[str] = None):
    query = {"user_id": user.user_id}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    dividends = await db.dividends.find(query, {"_id": 0}).to_list(10000)
    
    by_month = {}
    by_ticker = {}
    
    for d in dividends:
        month = d["payment_date"][:7]
        by_month[month] = by_month.get(month, 0) + d["amount"]
        by_ticker[d["ticker"]] = by_ticker.get(d["ticker"], 0) + d["amount"]
    
    return {
        "total": round(sum(d["amount"] for d in dividends), 2),
        "by_month": [{"month": k, "amount": round(v, 2)} for k, v in sorted(by_month.items())],
        "by_ticker": [{"ticker": k, "amount": round(v, 2)} for k, v in sorted(by_ticker.items(), key=lambda x: -x[1])]
    }

@api_router.delete("/dividends/all")
async def delete_all_dividends(user: User = Depends(get_current_user), portfolio_id: Optional[str] = None):
    """Delete all dividends for the current user, optionally filtered by portfolio"""
    query = {"user_id": user.user_id}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    result = await db.dividends.delete_many(query)
    return {"message": f"{result.deleted_count} dividendos excluídos", "deleted": result.deleted_count}

@api_router.post("/dividends/sync")
async def sync_dividends(user: User = Depends(get_current_user), portfolio_id: Optional[str] = None):
    """
    Synchronize dividends from Investidor10 for all stocks in the user's portfolio.
    This checks if the user held the stock on the 'data com' (ex-date) and creates dividend entries.
    """
    query = {"user_id": user.user_id}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    stocks = await db.stocks.find(query, {"_id": 0}).to_list(1000)
    
    if not stocks:
        return {"message": "Nenhuma ação na carteira", "synced": 0, "total_tickers": 0}
    
    # Get unique tickers
    unique_tickers = list(set(s["ticker"] for s in stocks))
    
    synced = 0
    skipped = 0
    errors = []
    details = []
    
    for ticker in unique_tickers:
        try:
            # Fetch dividends from Investidor10
            dividends_data = fetch_investidor10_dividends(ticker)
            
            if not dividends_data:
                logger.info(f"No dividends found for {ticker}")
                continue
            
            # Get all user stocks for this ticker (there might be multiple purchases)
            user_stocks = [s for s in stocks if s["ticker"] == ticker]
            
            for div_data in dividends_data:
                data_com = div_data.get("data_com")
                data_pagamento = div_data.get("data_pagamento") or data_com
                valor = div_data.get("valor", 0)
                tipo = div_data.get("tipo", "Dividendo").lower()
                
                if not data_com or valor <= 0:
                    continue
                
                # Check if user held the stock on data_com for each purchase
                for stock in user_stocks:
                    purchase_date = stock.get("purchase_date")
                    
                    # If no purchase date is set, skip this dividend (we can't verify eligibility)
                    if not purchase_date:
                        logger.info(f"Skipping dividend for {ticker}: no purchase_date set for stock")
                        skipped += 1
                        continue
                    
                    try:
                        # Parse dates for comparison
                        if isinstance(purchase_date, str):
                            purchase_dt = datetime.strptime(purchase_date, "%Y-%m-%d").date()
                        else:
                            purchase_dt = purchase_date
                        
                        data_com_dt = datetime.strptime(data_com, "%Y-%m-%d").date()
                        
                        # User must have purchased BEFORE or ON the data_com to be eligible
                        # Stocks bought AFTER the data_com do NOT receive the dividend
                        user_held_on_data_com = purchase_dt <= data_com_dt
                        
                    except Exception as e:
                        logger.error(f"Date parsing error for {ticker}: {e}")
                        skipped += 1
                        continue
                    
                    if not user_held_on_data_com:
                        logger.debug(f"Skipping dividend for {ticker}: purchased {purchase_date} > data_com {data_com}")
                        skipped += 1
                        continue
                    
                    # Calculate total dividend amount based on quantity
                    quantity = stock.get("quantity", 0)
                    total_dividend = round(valor * quantity, 2)
                    
                    if total_dividend <= 0:
                        continue
                    
                    # Check if this dividend already exists (avoid duplicates)
                    existing = await db.dividends.find_one({
                        "user_id": user.user_id,
                        "ticker": ticker,
                        "payment_date": data_pagamento,
                        "amount": total_dividend
                    })
                    
                    if existing:
                        skipped += 1
                        continue
                    
                    # Create dividend entry
                    dividend = Dividend(
                        user_id=user.user_id,
                        stock_id=stock["stock_id"],
                        ticker=ticker,
                        portfolio_id=stock.get("portfolio_id"),  # Include portfolio_id from stock
                        amount=total_dividend,
                        payment_date=data_pagamento,
                        type="jcp" if "jcp" in tipo or "juros" in tipo else "dividendo"
                    )
                    doc = dividend.model_dump()
                    doc["created_at"] = doc["created_at"].isoformat()
                    doc["ex_date"] = data_com  # Store the ex-date for reference
                    await db.dividends.insert_one(doc)
                    
                    synced += 1
                    details.append({
                        "ticker": ticker,
                        "amount": total_dividend,
                        "payment_date": data_pagamento,
                        "ex_date": data_com
                    })
                    
                    logger.info(f"Synced dividend for {ticker}: R${total_dividend} (ex-date: {data_com})")
                    
        except Exception as e:
            logger.error(f"Error syncing dividends for {ticker}: {e}")
            errors.append({"ticker": ticker, "error": str(e)})
    
    result = {
        "message": f"Sincronização concluída: {synced} dividendos adicionados",
        "synced": synced,
        "skipped": skipped,
        "total_tickers": len(unique_tickers),
        "details": details[:10]  # Return first 10 details
    }
    
    if errors:
        result["errors"] = errors
    
    return result

# ==================== VALUATION ROUTES ====================

@api_router.post("/valuation/calculate")
async def calculate_valuation(data: ValuationRequest, user: User = Depends(get_current_user)):
    results = {}
    
    # Modelo de Gordon
    if data.discount_rate > data.dividend_growth_rate:
        gordon_price = data.dividend_per_share * (1 + data.dividend_growth_rate) / (data.discount_rate - data.dividend_growth_rate)
        results["gordon"] = {
            "name": "Modelo de Gordon",
            "ceiling_price": round(gordon_price, 2),
            "upside": round((gordon_price / data.current_price - 1) * 100, 2),
            "recommendation": "Comprar" if gordon_price > data.current_price else "Aguardar"
        }
    
    # Método Bazin
    if data.desired_yield > 0:
        bazin_price = data.dividend_per_share / data.desired_yield
        results["bazin"] = {
            "name": "Método Bazin",
            "ceiling_price": round(bazin_price, 2),
            "upside": round((bazin_price / data.current_price - 1) * 100, 2),
            "recommendation": "Comprar" if bazin_price > data.current_price else "Aguardar"
        }
    
    # DCF Simplificado
    if data.free_cash_flow and data.shares_outstanding and data.discount_rate > data.growth_rate:
        terminal_value = data.free_cash_flow * (1 + data.growth_rate) / (data.discount_rate - data.growth_rate)
        dcf_price = terminal_value / data.shares_outstanding
        results["dcf"] = {
            "name": "Fluxo de Caixa Descontado",
            "ceiling_price": round(dcf_price, 2),
            "upside": round((dcf_price / data.current_price - 1) * 100, 2),
            "recommendation": "Comprar" if dcf_price > data.current_price else "Aguardar"
        }
    
    # Método Warren Buffett (Owner Earnings com DCF projetado)
    # 1. Calcular Owner Earnings = Lucro Líquido + Depreciação - CapEx
    # 2. Taxa de Crescimento = ROE × (1 - Payout)
    # 3. Projetar FCF para 10 anos com a taxa de crescimento
    # 4. Descontar os fluxos usando taxa de retorno mínima
    # 5. Calcular valor terminal
    # 6. Somar valores presentes e dividir por número de ações
    if data.net_income and data.depreciation is not None and data.capex is not None and data.shares_outstanding:
        owner_earnings = data.net_income + data.depreciation - data.capex
        
        # Calcular taxa de crescimento baseada em ROE e Payout
        # Growth = ROE × (1 - Payout/100)
        if data.roe and data.payout:
            retention_rate = 1 - (data.payout / 100)
            growth_rate = (data.roe / 100) * retention_rate
        else:
            growth_rate = data.growth_rate or 0.05
        
        # Limitar crescimento a valores razoáveis
        growth_rate = min(max(growth_rate, 0), 0.25)  # 0% a 25%
        
        discount_rate = data.discount_rate  # Taxa de retorno mínimo esperado
        
        # Projeção de 10 anos de Owner Earnings
        projection_years = 10
        projected_cash_flows = []
        current_fcf = owner_earnings
        
        for year in range(1, projection_years + 1):
            projected_fcf = current_fcf * ((1 + growth_rate) ** year)
            present_value = projected_fcf / ((1 + discount_rate) ** year)
            projected_cash_flows.append({
                "year": year,
                "fcf": round(projected_fcf, 0),
                "present_value": round(present_value, 0)
            })
        
        # Somar valores presentes dos fluxos projetados
        sum_pv = sum(cf["present_value"] for cf in projected_cash_flows)
        
        # Valor Terminal (perpetuidade com crescimento de 3% após ano 10)
        terminal_growth = 0.03  # Crescimento na perpetuidade
        if discount_rate > terminal_growth:
            terminal_fcf = projected_cash_flows[-1]["fcf"] * (1 + terminal_growth)
            terminal_value = terminal_fcf / (discount_rate - terminal_growth)
            terminal_pv = terminal_value / ((1 + discount_rate) ** projection_years)
        else:
            terminal_pv = 0
        
        # Valor Intrínseco Total
        intrinsic_value = sum_pv + terminal_pv
        
        # Preço por ação (Valor Intrínseco)
        buffett_price_raw = intrinsic_value / data.shares_outstanding
        
        # Aplicar margem de segurança de 25% (comprar com desconto)
        margin_of_safety = 0.25
        buffett_price = buffett_price_raw * (1 - margin_of_safety)
        
        results["buffett"] = {
            "name": "Método Warren Buffett",
            "description": "DCF com Owner Earnings (10 anos + perptuidade)",
            "owner_earnings": round(owner_earnings / 1000000, 2),  # Em milhões
            "growth_rate_used": f"{growth_rate * 100:.1f}%",
            "discount_rate": f"{discount_rate * 100:.1f}%",
            "roe": f"{data.roe:.1f}%" if data.roe else "N/A",
            "payout": f"{data.payout:.1f}%" if data.payout else "N/A",
            "sum_projected_pv": round(sum_pv / 1000000, 2),  # Em milhões
            "terminal_value_pv": round(terminal_pv / 1000000, 2),  # Em milhões
            "intrinsic_value_total": round(intrinsic_value / 1000000, 2),  # Em milhões
            "intrinsic_price_per_share": round(buffett_price_raw, 2),
            "ceiling_price": round(buffett_price, 2),
            "margin_of_safety": f"{margin_of_safety * 100:.0f}%",
            "upside": round((buffett_price / data.current_price - 1) * 100, 2),
            "recommendation": "Comprar" if buffett_price > data.current_price else "Aguardar"
        }
    
    return {
        "ticker": data.ticker,
        "current_price": data.current_price,
        "valuations": results
    }

# ==================== AI ANALYSIS ROUTES ====================

@api_router.post("/analysis/stock")
async def analyze_stock(data: AnalysisRequest, user: User = Depends(get_current_user)):
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="LLM key not configured")
        
        prompt = f"""Analise a ação {data.ticker} da bolsa brasileira B3.

Dados disponíveis:
- Preço atual: R$ {data.current_price}
- Setor: {data.sector or 'Não informado'}
- Dividend Yield: {data.dividend_yield or 'Não informado'}%
- P/L: {data.pe_ratio or 'Não informado'}

{f'Pergunta específica: {data.question}' if data.question else ''}

Faça uma análise fundamentalista breve e objetiva em português, incluindo:
1. Avaliação do preço atual
2. Pontos fortes e fracos
3. Recomendação (Comprar, Manter ou Vender)

Seja direto e use no máximo 200 palavras."""
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"analysis_{user.user_id}_{data.ticker}",
            system_message="Você é um analista financeiro especializado no mercado brasileiro de ações. Forneça análises precisas e objetivas."
        ).with_model("openai", "gpt-4o-mini")
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        return {
            "ticker": data.ticker,
            "analysis": response,
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
    except ImportError:
        raise HTTPException(status_code=500, detail="LLM integration not available")
    except Exception as e:
        logger.error(f"Analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/analysis/portfolio")
async def analyze_portfolio(user: User = Depends(get_current_user)):
    """Analyze the entire portfolio using AI"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="LLM key not configured")
        
        # Get user's stocks
        stocks = await db.stocks.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
        dividends = await db.dividends.find({"user_id": user.user_id}, {"_id": 0}).to_list(10000)
        
        if not stocks:
            raise HTTPException(status_code=400, detail="Nenhuma ação na carteira para analisar")
        
        # Aggregate stocks by ticker
        portfolio_summary = {}
        total_invested = 0
        total_current = 0
        
        for stock in stocks:
            ticker = stock["ticker"]
            invested = stock["quantity"] * stock["average_price"]
            current = stock["quantity"] * (stock.get("current_price") or stock["average_price"])
            
            if ticker not in portfolio_summary:
                portfolio_summary[ticker] = {
                    "ticker": ticker,
                    "name": stock.get("name", ticker),
                    "sector": stock.get("sector", "Não informado"),
                    "quantity": 0,
                    "invested": 0,
                    "current": 0,
                    "current_price": stock.get("current_price"),
                    "average_price": 0,
                }
            
            portfolio_summary[ticker]["quantity"] += stock["quantity"]
            portfolio_summary[ticker]["invested"] += invested
            portfolio_summary[ticker]["current"] += current
            total_invested += invested
            total_current += current
        
        # Calculate averages and percentages
        for ticker, data in portfolio_summary.items():
            data["average_price"] = data["invested"] / data["quantity"] if data["quantity"] > 0 else 0
            data["gain_percent"] = ((data["current"] / data["invested"]) - 1) * 100 if data["invested"] > 0 else 0
            data["portfolio_percent"] = (data["current"] / total_current) * 100 if total_current > 0 else 0
        
        # Calculate dividends by ticker
        dividends_by_ticker = {}
        total_dividends = 0
        for div in dividends:
            ticker = div["ticker"]
            dividends_by_ticker[ticker] = dividends_by_ticker.get(ticker, 0) + div["amount"]
            total_dividends += div["amount"]
        
        # Build portfolio description
        portfolio_desc = []
        for ticker, data in sorted(portfolio_summary.items(), key=lambda x: -x[1]["current"]):
            div_received = dividends_by_ticker.get(ticker, 0)
            portfolio_desc.append(
                f"- {ticker} ({data['sector']}): {data['quantity']} ações, "
                f"PM R${data['average_price']:.2f}, Atual R${data.get('current_price', 0):.2f}, "
                f"Rendimento {data['gain_percent']:+.1f}%, "
                f"Peso {data['portfolio_percent']:.1f}%, "
                f"Dividendos recebidos R${div_received:.2f}"
            )
        
        total_gain = total_current - total_invested
        total_gain_percent = ((total_current / total_invested) - 1) * 100 if total_invested > 0 else 0
        total_return = total_gain + total_dividends
        total_return_percent = (total_return / total_invested) * 100 if total_invested > 0 else 0
        
        prompt = f"""Analise esta carteira de ações brasileiras de forma completa e estratégica:

RESUMO DA CARTEIRA:
- Total Investido: R$ {total_invested:,.2f}
- Valor Atual: R$ {total_current:,.2f}
- Ganho de Capital: R$ {total_gain:,.2f} ({total_gain_percent:+.1f}%)
- Dividendos Recebidos: R$ {total_dividends:,.2f}
- Retorno Total: R$ {total_return:,.2f} ({total_return_percent:+.1f}%)
- Número de Ativos: {len(portfolio_summary)}

COMPOSIÇÃO:
{chr(10).join(portfolio_desc)}

Por favor, forneça uma análise completa em português incluindo:

1. **DIVERSIFICAÇÃO**: Avalie a diversificação setorial e de ativos. A carteira está bem diversificada?

2. **CONCENTRAÇÃO**: Identifique riscos de concentração excessiva em algum ativo ou setor.

3. **QUALIDADE DOS ATIVOS**: Comente sobre a qualidade geral das empresas na carteira.

4. **PONTOS FORTES**: Quais são os pontos fortes desta carteira?

5. **PONTOS DE ATENÇÃO**: Quais aspectos merecem atenção ou podem ser melhorados?

6. **SUGESTÕES**: Dê 2-3 sugestões práticas para otimizar a carteira.

7. **NOTA GERAL**: Dê uma nota de 0 a 10 para a carteira e justifique brevemente.

Seja objetivo e direto, use linguagem acessível."""

        chat = LlmChat(
            api_key=api_key,
            session_id=f"portfolio_analysis_{user.user_id}",
            system_message="Você é um consultor financeiro especializado em análise de carteiras de ações brasileiras. Forneça análises estratégicas, objetivas e acionáveis."
        ).with_model("openai", "gpt-4o-mini")
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        return {
            "analysis": response,
            "summary": {
                "total_invested": round(total_invested, 2),
                "total_current": round(total_current, 2),
                "total_gain": round(total_gain, 2),
                "total_gain_percent": round(total_gain_percent, 2),
                "total_dividends": round(total_dividends, 2),
                "total_return": round(total_return, 2),
                "total_return_percent": round(total_return_percent, 2),
                "stocks_count": len(portfolio_summary),
            },
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
    except ImportError:
        raise HTTPException(status_code=500, detail="LLM integration not available")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Portfolio analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/")
async def root():
    return {"message": "Stock Portfolio API"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
